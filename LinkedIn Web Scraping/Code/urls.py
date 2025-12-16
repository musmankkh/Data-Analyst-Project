import random
import time
import os
import pickle
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook

# ---------------- RANDOM WAIT FUNCTION ----------------
def human_sleep():
    t = random.uniform(6, 12)
    print(f"⏳ Human sleep: {round(t, 2)} sec")
    time.sleep(t)

# ---------------- CONFIG ----------------
EMAIL = "usman12kh46@gmail.com"
PASSWORD = "Usman46@@"
SEARCH_QUERY = '"CTO"' 

EXCEL_FILE = "C:\\Users\\Usman Asghar\\Desktop\\Intern\\LinkedIn Web Scraping\\Excel File\\linkedin_profiles.xlsx"
COOKIES_FILE = "C:\\Users\\Usman Asghar\\Desktop\\Intern\\LinkedIn Web Scraping\\linkedins_cookies.pkl"

# ---------------- LOAD EXISTING EXCEL (IF ANY) ----------------
existing_urls = set()
if os.path.exists(EXCEL_FILE):
    print("📂 Existing Excel file found. Loading URLs...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[0]
        if url:
            existing_urls.add(url)
    print(f"✅ Loaded {len(existing_urls)} existing URLs from Excel.")
else:
    print("📄 No existing Excel file found. A new one will be created.")
    wb = Workbook()
    ws = wb.active
    ws.title = "LinkedIn Profiles"
    ws.append(["Profile URL"])

# ---------------- DYNAMIC TARGET ----------------
TARGET_COUNT = 2000 - len(existing_urls)
print(f"📌 Need {TARGET_COUNT} new URLs to reach 2000 total.")
if TARGET_COUNT <= 0:
    print("🎉 Already have 2000+ URLs in Excel. No scraping needed.")
    exit()

# ---------------- BROWSER SETUP ----------------
options = Options()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 50)

# ---------------- LOGIN WITH COOKIE MANAGEMENT ----------------
driver.get("https://www.linkedin.com")

# Try to load cookies if they exist
if os.path.exists(COOKIES_FILE):
    print("🍪 Found saved cookies. Loading...")
    with open(COOKIES_FILE, "rb") as file:
        cookies = pickle.load(file)
        for cookie in cookies:
            driver.add_cookie(cookie)
    
    # Refresh to apply cookies
    driver.refresh()
    time.sleep(3)
    
    # Check if login was successful
    try:
        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[contains(@placeholder,'Search')]")
        ))
        print("✅ Successfully logged in using saved cookies!")
    except:
        print("⚠️ Cookies expired or invalid. Logging in manually...")
        os.remove(COOKIES_FILE)  # Remove invalid cookies
        driver.get("https://www.linkedin.com/login")
        
        wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(EMAIL)
        driver.find_element(By.ID, "password").send_keys(PASSWORD)
        driver.find_element(By.XPATH, "//button[@type='submit']").click()
        time.sleep(5)
        
        # Save new cookies
        print("💾 Saving new cookies...")
        with open(COOKIES_FILE, "wb") as file:
            pickle.dump(driver.get_cookies(), file)
        print("✅ Cookies saved successfully!")
else:
    print("🔐 No saved cookies found. Logging in...")
    driver.get("https://www.linkedin.com/login")
    
    wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(EMAIL)
    driver.find_element(By.ID, "password").send_keys(PASSWORD)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(5)
    
    # Save cookies after successful login
    print("💾 Saving cookies for future use...")
    with open(COOKIES_FILE, "wb") as file:
        pickle.dump(driver.get_cookies(), file)
    print("✅ Cookies saved successfully!")

# ---------------- SEARCH ----------------
search_box = wait.until(EC.presence_of_element_located(
    (By.XPATH, "//input[contains(@placeholder,'Search')]")
))
search_box.send_keys(SEARCH_QUERY)
search_box.send_keys(Keys.RETURN)

# ---------------- CLICK "SEE ALL PEOPLE RESULTS" ----------------
see_all = wait.until(EC.element_to_be_clickable(
    (By.XPATH, "//a[contains(.,'See all people results')]")
))
driver.execute_script("arguments[0].click();", see_all)
human_sleep()

# ---------------- SCRAPING LOOP ----------------
profile_urls = set()
page_number = 1

while len(profile_urls) < TARGET_COUNT:
    print(f"\n📌 Scraping Page {page_number}...")

    # scroll (human style)
    for _ in range(8):  # scroll more for better results
        driver.execute_script("window.scrollBy(0, 2000);")
        time.sleep(random.uniform(2.5, 4.5))

    # find all profile links
    links = driver.find_elements(By.XPATH, "//a[contains(@href, '/in/')]")
    for link in links:
        url = link.get_attribute("href")
        if url and "/in/" in url and url not in existing_urls and url not in profile_urls:
            profile_urls.add(url)
            print(f"✔ New collected: {url}")
            if len(profile_urls) >= TARGET_COUNT:
                break

    print(f"➡ New collected so far in this run: {len(profile_urls)}")

    if len(profile_urls) >= TARGET_COUNT:
        break

    # CLICK NEXT BUTTON
    try:
        next_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[.//span[text()='Next'] or .//span[contains(text(),'Next')]]")
        ))
        driver.execute_script("arguments[0].click();", next_btn)
        print("➡ Clicked Next Page")
        human_sleep()
        page_number += 1
    except Exception:
        print("❌ No more pages or Next button not found.")
        break

driver.quit()

# ---------------- SAVE TO EXCEL ----------------
print("\n💾 Saving results to Excel...")
for url in profile_urls:
    ws.append([url])

wb.save(EXCEL_FILE)

print(f"🎉 Scraping complete! File saved as: {EXCEL_FILE}")
print(f"➕ New URLs added this run: {len(profile_urls)}")
print(f"📊 Total URLs now (old + new): {len(existing_urls) + len(profile_urls)}")